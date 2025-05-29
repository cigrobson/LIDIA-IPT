from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import os
import json
from datetime import datetime, timedelta
import uuid
import docx
import docx2txt
from werkzeug.utils import secure_filename
import io
import csv
import openpyxl
import chardet
import re
import time

# Imports para Supabase
try:
    from supabase import create_client, Client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False

# Imports condicionais para bibliotecas robustas
try:
    import fitz  # pymupdf
    PYMUPDF_AVAILABLE = True
except ImportError:
    import PyPDF2
    PYMUPDF_AVAILABLE = False

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'lidia-ipt-secret-key-2024')

# Configurações Supabase
SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_KEY = os.environ.get('SUPABASE_KEY', '')
ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')

# Cliente Supabase
supabase: Client = None
if SUPABASE_AVAILABLE and SUPABASE_URL and SUPABASE_KEY:
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY, options=None)
        print("✅ Supabase conectado com sucesso!")
    except Exception as e:
        print(f"❌ Erro ao conectar Supabase: {e}")
        supabase = None
else:
    print("❌ Supabase não configurado - usando modo fallback")

class AdvancedDocumentProcessor:
    """Processador de documentos avançado com chunking inteligente e fallbacks"""
    
    def __init__(self):
        self.max_chunk_size = 1000  # palavras por chunk
        self.chunk_overlap = 100    # palavras de sobreposição
        self.max_context_length = 4000  # caracteres totais
        
        # Verificar disponibilidade de bibliotecas avançadas
        self.pymupdf_available = PYMUPDF_AVAILABLE
        self.docx2txt_available = self._check_docx2txt()
        self.chardet_available = self._check_chardet()
        
        print(f"[RAG] Inicializando processador avançado:")
        print(f"[RAG] - PyMuPDF: {'✅' if self.pymupdf_available else '❌'}")
        print(f"[RAG] - docx2txt: {'✅' if self.docx2txt_available else '❌'}")
        print(f"[RAG] - chardet: {'✅' if self.chardet_available else '❌'}")
        print(f"[RAG] - Supabase: {'✅' if supabase else '❌'}")
    
    def _check_docx2txt(self):
        try:
            import docx2txt
            return True
        except ImportError:
            return False
    
    def _check_chardet(self):
        try:
            import chardet
            return True
        except ImportError:
            return False
    
    @staticmethod
    def extract_text_from_file(file_storage):
        """Extrai texto com melhorias avançadas"""
        processor = AdvancedDocumentProcessor()
        return processor._extract_with_fallbacks(file_storage)
    
    def _extract_with_fallbacks(self, file_storage):
        """Sistema híbrido com fallbacks seguros"""
        try:
            filename = file_storage.filename.lower()
            file_content = file_storage.read()
            
            print(f"[RAG] Processando: {filename} ({len(file_content)} bytes)")
            
            # Determinar encoding se necessário
            encoding = self._detect_encoding_safe(file_content) if filename.endswith(('.txt', '.csv')) else 'utf-8'
            
            # Estratégia baseada no tipo de arquivo
            if filename.endswith('.pdf'):
                text = self._extract_pdf_improved(file_content)
            elif filename.endswith('.docx'):
                text = self._extract_docx_improved(file_content)
            elif filename.endswith('.txt'):
                text = self._extract_text_safe(file_content, encoding)
            elif filename.endswith('.csv'):
                text = self._extract_csv_improved(file_content, encoding)
            elif filename.endswith('.xlsx'):
                text = self._extract_xlsx_improved(file_content)
            else:
                return "Formato de arquivo não suportado para extração de texto."
            
            if not text or len(text.strip()) < 10:
                return "Não foi possível extrair conteúdo significativo do arquivo."
            
            # Aplicar chunking inteligente
            processed_text = self._intelligent_chunking(text, filename)
            
            print(f"[RAG] Sucesso: {len(processed_text)} caracteres extraídos")
            return processed_text
            
        except Exception as e:
            print(f"[RAG] Erro: {str(e)}")
            return f"Erro ao processar arquivo: {str(e)}"
    
    def _detect_encoding_safe(self, file_content):
        """Detecta encoding com fallback seguro"""
        if self.chardet_available:
            try:
                import chardet
                result = chardet.detect(file_content)
                detected = result.get('encoding', 'utf-8')
                print(f"[RAG] Encoding detectado: {detected}")
                return detected
            except:
                pass
        return 'utf-8'
    
    def _extract_pdf_improved(self, file_content):
        """PDF com PyMuPDF se disponível, senão PyPDF2"""
        text = ""
        
        # Tentativa 1: PyMuPDF (se disponível)
        if self.pymupdf_available:
            try:
                import fitz
                pdf_document = fitz.open(stream=file_content, filetype="pdf")
                for page_num in range(min(pdf_document.page_count, 100)):
                    page = pdf_document[page_num]
                    page_text = page.get_text()
                    if page_text.strip():
                        text += page_text + "\n"
                pdf_document.close()
                
                if text.strip():
                    print("[RAG] PDF extraído com PyMuPDF")
                    return text
            except Exception as e:
                print(f"[RAG] PyMuPDF falhou: {e}")
        
        # Fallback: PyPDF2
        try:
            import PyPDF2
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
            for page in pdf_reader.pages[:100]:
                page_text = page.extract_text()
                if page_text.strip():
                    text += page_text + "\n"
            
            if text.strip():
                print("[RAG] PDF extraído com PyPDF2")
                return text
        except Exception as e:
            print(f"[RAG] PyPDF2 falhou: {e}")
        
        return "Não foi possível extrair texto do PDF. Arquivo pode estar protegido ou corrompido."
    
    def _extract_docx_improved(self, file_content):
        """DOCX com docx2txt se disponível, senão python-docx"""
        text = ""
        
        # Tentativa 1: docx2txt (se disponível)
        if self.docx2txt_available:
            try:
                import docx2txt
                text = docx2txt.process(io.BytesIO(file_content))
                if text and text.strip():
                    print("[RAG] DOCX extraído com docx2txt")
                    return text
            except Exception as e:
                print(f"[RAG] docx2txt falhou: {e}")
        
        # Fallback: python-docx
        try:
            doc = docx.Document(io.BytesIO(file_content))
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            
            if text.strip():
                print("[RAG] DOCX extraído com python-docx")
                return text
        except Exception as e:
            print(f"[RAG] python-docx falhou: {e}")
        
        return "Erro ao processar documento Word."
    
    def _extract_text_safe(self, file_content, encoding):
        """Extração de texto com múltiplos encodings"""
        encodings_to_try = [encoding, 'utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        
        for enc in encodings_to_try:
            try:
                text = file_content.decode(enc, errors='replace')
                if text and len(text.strip()) > 0:
                    print(f"[RAG] Texto extraído com encoding: {enc}")
                    return text
            except:
                continue
        
        return "Erro ao decodificar arquivo de texto."
    
    def _extract_csv_improved(self, file_content, encoding):
        """CSV com melhor detecção de delimitador"""
        try:
            text = file_content.decode(encoding, errors='replace')
            
            # Detectar delimitador mais provável
            delimiters = [',', ';', '\t', '|']
            best_delimiter = ','
            max_columns = 0
            
            for delimiter in delimiters:
                try:
                    sample_reader = csv.reader(io.StringIO(text[:1000]), delimiter=delimiter)
                    first_row = next(sample_reader, [])
                    if len(first_row) > max_columns:
                        max_columns = len(first_row)
                        best_delimiter = delimiter
                except:
                    continue
            
            # Processar CSV com melhor delimitador
            csv_reader = csv.reader(io.StringIO(text), delimiter=best_delimiter)
            rows = list(csv_reader)
            
            if not rows:
                return "Arquivo CSV vazio."
            
            # Formato otimizado para análise
            result = f"DADOS CSV (delimitador: '{best_delimiter}'):\n"
            headers = rows[0] if rows else []
            result += f"Colunas ({len(headers)}): {', '.join(headers[:10])}\n"
            result += f"Total de registros: {len(rows)-1}\n\n"
            
            # Amostra dos dados
            sample_size = min(20, len(rows))
            result += "AMOSTRA DOS DADOS:\n"
            
            for i, row in enumerate(rows[:sample_size]):
                if i == 0:  # Header
                    result += f"CABEÇALHO: {' | '.join(str(cell)[:40] for cell in row[:8])}\n"
                else:
                    clean_row = [str(cell)[:40] for cell in row[:8]]
                    result += f"Reg {i}: {' | '.join(clean_row)}\n"
            
            if len(rows) > sample_size:
                result += f"\n... e mais {len(rows)-sample_size} registros"
            
            print(f"[RAG] CSV processado: {len(rows)} linhas, delimitador '{best_delimiter}'")
            return result
            
        except Exception as e:
            print(f"[RAG] Erro no CSV: {e}")
            return f"Erro ao processar CSV: {str(e)}"
    
    def _extract_xlsx_improved(self, file_content):
        """XLSX com melhor tratamento"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            result = "DADOS EXCEL:\n"
            
            for sheet_name in workbook.sheetnames[:5]:  # Máximo 5 planilhas
                sheet = workbook[sheet_name]
                result += f"\n=== PLANILHA: {sheet_name} ===\n"
                
                rows = list(sheet.iter_rows(values_only=True, max_row=50))
                if rows:
                    non_empty_rows = [row for row in rows if any(cell is not None for cell in row)]
                    
                    if non_empty_rows:
                        result += f"Dimensões: {len(non_empty_rows)} linhas úteis\n"
                        
                        # Header e amostra
                        if non_empty_rows:
                            header = [str(cell)[:30] if cell is not None else "" for cell in non_empty_rows[0][:8]]
                            result += f"Colunas: {' | '.join(header)}\n\n"
                        
                        for i, row in enumerate(non_empty_rows[:12]):
                            clean_row = [str(cell)[:30] if cell is not None else "" for cell in row[:8]]
                            result += f"Linha {i+1}: {' | '.join(clean_row)}\n"
                        
                        if len(non_empty_rows) > 12:
                            result += f"... e mais {len(non_empty_rows)-12} linhas\n"
            
            print(f"[RAG] Excel processado com {len(workbook.sheetnames)} planilhas")
            return result
            
        except Exception as e:
            return f"Erro ao processar Excel: {str(e)}"
    
    def _intelligent_chunking(self, text, filename):
        """Chunking inteligente do texto baseado em contexto"""
        try:
            # Limpar e normalizar texto
            text = self._clean_text_safe(text)
            
            if len(text) <= self.max_context_length:
                return text
            
            # Dividir em chunks inteligentes mantendo contexto
            chunks = self._create_smart_chunks(text)
            
            # Combinar melhores chunks
            if chunks:
                return self._combine_best_chunks(chunks, filename)
            else:
                return text[:self.max_context_length]
                
        except Exception as e:
            print(f"[RAG] Erro no chunking: {e}")
            return text[:self.max_context_length]
    
    def _create_smart_chunks(self, text):
        """Cria chunks inteligentes com sobreposição"""
        sentences = re.split(r'[.!?]+', text)
        chunks = []
        current_chunk = ""
        current_words = 0
        
        for sentence in sentences:
            sentence = sentence.strip()
            if not sentence:
                continue
            
            words_in_sentence = len(sentence.split())
            
            if current_words + words_in_sentence > self.max_chunk_size and current_chunk:
                chunks.append(current_chunk.strip())
                
                # Overlap com últimas palavras
                overlap_words = current_chunk.split()[-self.chunk_overlap:]
                current_chunk = " ".join(overlap_words) + " " + sentence
                current_words = len(current_chunk.split())
            else:
                current_chunk += " " + sentence
                current_words += words_in_sentence
        
        if current_chunk.strip():
            chunks.append(current_chunk.strip())
        
        return chunks
    
    def _combine_best_chunks(self, chunks, filename):
        """Combina os melhores chunks baseado no contexto"""
        if not chunks:
            return ""
        
        if len(chunks) <= 2:
            combined = "\n\n".join(chunks)
            return combined[:self.max_context_length]
        
        # Priorizar chunks com conteúdo mais relevante
        scored_chunks = []
        for i, chunk in enumerate(chunks):
            score = self._score_chunk_relevance(chunk, filename)
            scored_chunks.append((score, i, chunk))
        
        # Ordenar por relevância
        scored_chunks.sort(reverse=True)
        
        # Combinar melhores chunks até o limite
        result = ""
        for score, i, chunk in scored_chunks:
            if len(result) + len(chunk) + 10 <= self.max_context_length:
                if result:
                    result += "\n\n"
                result += f"[Seção {i+1}]\n{chunk}"
            else:
                break
        
        return result if result else chunks[0][:self.max_context_length]
    
    def _score_chunk_relevance(self, chunk, filename):
        """Pontua relevância do chunk"""
        score = 0
        chunk_lower = chunk.lower()
        
        # Pontuação por palavras-chave técnicas
        technical_terms = ['resultado', 'conclusão', 'objetivo', 'método', 'análise', 
                          'dados', 'tabela', 'gráfico', 'resumo', 'abstract', 'introdução']
        score += sum(chunk_lower.count(term) for term in technical_terms) * 10
        
        # Pontuação por estrutura
        if any(marker in chunk for marker in ['1.', '2.', 'a)', 'b)', '•', '-', 'I.', 'II.']):
            score += 8
        
        # Bonificar chunks com números/dados
        if re.search(r'\d+[%,$]|\d+\.\d+|\d+/\d+', chunk):
            score += 12
        
        # Penalizar chunks muito curtos
        if len(chunk.split()) < 30:
            score -= 10
        
        return score
    
    def _clean_text_safe(self, text):
        """Limpeza básica e segura do texto"""
        if not text:
            return ""
        
        # Remover quebras excessivas
        text = re.sub(r'\n\s*\n\s*\n+', '\n\n', text)
        # Remover espaços excessivos
        text = re.sub(r' +', ' ', text)
        # Remover caracteres de controle problemáticos
        text = ''.join(char for char in text if ord(char) >= 32 or char in '\n\t')
        
        return text.strip()

class SupabaseManager:
    """Gerenciador de dados usando Supabase"""
    
    def __init__(self):
        self.client = supabase
        self.available = supabase is not None
        print(f"[DB] Supabase Manager: {'✅ Ativo' if self.available else '❌ Indisponível'}")
    
    def validate_ipt_email(self, email):
        if not email:
            return False, "Email obrigatório"
        
        if not email.endswith('@ipt.br'):
            return False, "Acesso exclusivo para colaboradores do IPT"
        
        if len(email.split('@')[0]) < 3:
            return False, "Email inválido"
        
        return True, "Email válido"
    
    def is_admin(self, email):
        """Verifica se email é administrador ativo"""
        if not self.available:
            return email == 'robsonss@ipt.br'  # Fallback
        
        try:
            result = self.client.table('administrators').select('is_active').eq('email', email).eq('is_active', True).execute()
            is_admin = len(result.data) > 0
            print(f"[DB] Admin check para {email}: {is_admin}")
            return is_admin
        except Exception as e:
            print(f"[DB] Erro ao verificar admin: {e}")
            return email == 'robsonss@ipt.br'  # Fallback
    
    def add_admin(self, email, granted_by):
        """Adiciona novo administrador"""
        if not self.available:
            return False, "Sistema de banco indisponível"
        
        try:
            data = {
                'email': email,
                'granted_by': granted_by,
                'granted_at': datetime.now().isoformat(),
                'is_active': True,
                'permissions': 'full'
            }
            
            result = self.client.table('administrators').insert(data).execute()
            
            if result.data:
                return True, "Administrador adicionado com sucesso"
            else:
                return False, "Erro ao adicionar administrador"
                
        except Exception as e:
            error_msg = str(e)
            if 'duplicate' in error_msg.lower() or 'unique' in error_msg.lower():
                return False, "Este email já é administrador"
            return False, f"Erro ao adicionar administrador: {error_msg}"
    
    def remove_admin(self, email):
        """Remove administrador"""
        if email == 'robsonss@ipt.br':
            return False, "Não é possível remover o administrador principal"
        
        if not self.available:
            return False, "Sistema de banco indisponível"
        
        try:
            result = self.client.table('administrators').update({'is_active': False}).eq('email', email).execute()
            
            if result.data:
                return True, "Administrador removido com sucesso"
            else:
                return False, "Administrador não encontrado"
                
        except Exception as e:
            return False, f"Erro ao remover administrador: {str(e)}"
    
    def get_admins(self):
        """Lista todos os administradores"""
        if not self.available:
            return [{'email': 'robsonss@ipt.br', 'granted_by': 'system', 
                    'granted_at': datetime.now().isoformat(), 'is_active': True, 'permissions': 'full'}]
        
        try:
            result = self.client.table('administrators').select('*').order('granted_at', desc=True).execute()
            return result.data
        except Exception as e:
            print(f"[DB] Erro ao listar admins: {e}")
            return []
    
    def log_conversation(self, email, chat_id, question, response, cost=0.003, document_context="", response_time=1.0, model="claude-3-haiku", context_length=0, chunk_count=0):
        """Registra conversa"""
        if not self.available:
            print(f"[DB] Conversa não registrada - banco indisponível")
            return
        
        try:
            # Inserir log de conversa
            conv_data = {
                'email': email,
                'chat_id': chat_id,
                'question': question,
                'response': response[:1000],  # Limitar tamanho
                'timestamp': datetime.now().isoformat(),
                'session_id': session.get('session_id', 'unknown'),
                'cost': cost,
                'document_context': document_context[:500] if document_context else "",
                'response_time': response_time,
                'model_used': model,
                'context_length': context_length,
                'chunk_count': chunk_count
            }
            
            self.client.table('conversation_logs').insert(conv_data).execute()
            
            # Atualizar ou criar chat
            chat_data = {
                'chat_id': chat_id,
                'email': email,
                'title': question[:50] + "..." if len(question) > 50 else question,
                'updated_at': datetime.now().isoformat(),
                'has_document': bool(document_context)
            }
            
            # Tentar update primeiro, se falhar, insert
            try:
                self.client.table('chats').update(chat_data).eq('chat_id', chat_id).execute()
            except:
                chat_data['created_at'] = datetime.now().isoformat()
                self.client.table('chats').insert(chat_data).execute()
            
            print(f"[DB] Conversa registrada: {chat_id}")
            
        except Exception as e:
            print(f"[DB] Erro ao registrar conversa: {e}")
    
    def store_document(self, email, chat_id, filename, original_content, processed_text, processing_time=0.0, extraction_method="", chunk_count=0):
        """Armazena documento processado"""
        if not self.available:
            print(f"[DB] Documento não armazenado - banco indisponível")
            return False
        
        try:
            file_type = filename.split('.')[-1].lower() if '.' in filename else 'unknown'
            processing_status = "success" if len(processed_text) > 50 else "partial"
            
            doc_data = {
                'email': email,
                'filename': filename,
                'original_size': len(original_content),
                'processed_size': len(processed_text),
                'upload_time': datetime.now().isoformat(),
                'chat_id': chat_id,
                'file_content': processed_text[:8000],  # Limitar para performance
                'file_type': file_type,
                'processing_time': processing_time,
                'extraction_method': extraction_method,
                'chunk_count': chunk_count,
                'processing_status': processing_status
            }
            
            self.client.table('file_uploads').insert(doc_data).execute()
            
            # Atualizar chat para indicar documento
            self.client.table('chats').update({
                'has_document': True,
                'document_name': filename
            }).eq('chat_id', chat_id).execute()
            
            print(f"[DB] Documento armazenado: {filename}")
            return True
            
        except Exception as e:
            print(f"[DB] Erro ao armazenar documento: {e}")
            return False
    
    def get_document_context(self, chat_id):
        """Recupera contexto do documento"""
        if not self.available:
            return None, None, 0
        
        try:
            result = self.client.table('file_uploads').select('file_content, filename, chunk_count').eq('chat_id', chat_id).order('upload_time', desc=True).limit(1).execute()
            
            if result.data:
                doc = result.data[0]
                return doc['file_content'], doc['filename'], doc.get('chunk_count', 0)
            
            return None, None, 0
            
        except Exception as e:
            print(f"[DB] Erro ao recuperar contexto: {e}")
            return None, None, 0
    
    def get_user_chats(self, email):
        """Lista chats do usuário"""
        if not self.available:
            return []
        
        try:
            result = self.client.table('chats').select('*').eq('email', email).order('updated_at', desc=True).limit(20).execute()
            return result.data
        except Exception as e:
            print(f"[DB] Erro ao listar chats: {e}")
            return []
    
    def get_chat_messages(self, email, chat_id):
        """Lista mensagens do chat"""
        if not self.available:
            return []
        
        try:
            result = self.client.table('conversation_logs').select('question, response, timestamp').eq('email', email).eq('chat_id', chat_id).order('timestamp', desc=False).execute()
            
            messages = []
            for msg in result.data:
                messages.append({'content': msg['question'], 'sender': 'user', 'timestamp': msg['timestamp']})
                messages.append({'content': msg['response'], 'sender': 'assistant', 'timestamp': msg['timestamp']})
            
            return messages
        except Exception as e:
            print(f"[DB] Erro ao listar mensagens: {e}")
            return []
    
    def log_access(self, email, ip_address, action="login", success=True, user_agent=""):
        """Registra acesso"""
        if not self.available:
            return
        
        try:
            access_data = {
                'email': email,
                'ip_address': ip_address,
                'timestamp': datetime.now().isoformat(),
                'action': action,
                'success': success,
                'user_agent': user_agent[:200]
            }
            
            self.client.table('access_logs').insert(access_data).execute()
            print(f"[DB] Acesso registrado: {email} - {action}")
            
        except Exception as e:
            print(f"[DB] Erro ao registrar acesso: {e}")
    
    def get_admin_stats(self):
        """Estatísticas para admin"""
        if not self.available:
            return {}
        
        try:
            stats = {}
            
            # Total de usuários únicos
            result = self.client.rpc('get_unique_users_count').execute()
            stats['total_users'] = result.data if result.data else 0
            
            # Usuários ativos (30 dias)
            thirty_days_ago = (datetime.now() - timedelta(days=30)).isoformat()
            result = self.client.table('conversation_logs').select('email', count='exact').gte('timestamp', thirty_days_ago).execute()
            stats['active_users_30d'] = result.count if result.count else 0
            
            # Total de conversas
            result = self.client.table('chats').select('chat_id', count='exact').execute()
            stats['total_conversations'] = result.count if result.count else 0
            
            # Conversas hoje
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
            result = self.client.table('conversation_logs').select('chat_id', count='exact').gte('timestamp', today).execute()
            stats['conversations_today'] = result.count if result.count else 0
            
            # Top usuários (simplificado)
            stats['top_users'] = [
                {'email': 'usuario@ipt.br', 'count': 10},
                {'email': 'outro@ipt.br', 'count': 8}
            ]
            
            # Atividade por dia (simplificado)
            stats['activity_by_day'] = [
                {'date': datetime.now().date().isoformat(), 'count': stats['conversations_today']}
            ]
            
            # Documentos por tipo
            result = self.client.table('file_uploads').select('file_type').execute()
            doc_types = {}
            for doc in result.data:
                file_type = doc['file_type']
                doc_types[file_type] = doc_types.get(file_type, 0) + 1
            
            stats['documents_by_type'] = [{'type': k, 'count': v} for k, v in doc_types.items()]
            
            return stats
            
        except Exception as e:
            print(f"[DB] Erro ao obter estatísticas: {e}")
            return {}

class LIDIAAssistant:
    def __init__(self):
        self.model = "claude-3-haiku-20240307"
        self.client = None
        
        self.ipt_context = """Você é LIDIA, a Inteligência Artificial do Laboratório Virtual de Inteligência Aplicada (LID) do Instituto de Pesquisas Tecnológicas (IPT).

Como assistente conversacional avançada com sistema RAG (Retrieval-Augmented Generation), você pode ajudar colaboradores do IPT com:
- Auxílio em pesquisas
- Suporte à redação de relatórios e documentação técnica
- Orientação sobre tecnologias e ferramentas disponíveis
- Interpretação e debates de resultados de pesquisa
- Sugestões de melhores práticas em projetos
- Resposta a dúvidas técnicas e conceituais
- Análise de documentos enviados pelos usuários

IMPORTANTE: Apenas responda sobre o que, quem você é ou sua finalidade se usuário perguntar especificamente. Nesse caso, use: "Eu sou a LIDIA, Inteligência Artificial do LID-IPT e estou sendo configurada para apoiar os colaborades do IPT".

Responda de forma cordial, clara, objetiva e profissional, sempre baseando suas respostas no contexto dos documentos quando disponível."""
    
    def get_client(self):
        if self.client is None:
            try:
                import anthropic
                if not ANTHROPIC_API_KEY or ANTHROPIC_API_KEY == '':
                    return None
                self.client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
            except Exception as e:
                print(f"Erro ao inicializar Anthropic: {e}")
                return None
        return self.client
    
    def process_query(self, message, user_email="", document_context="", filename="", chunk_count=0):
        start_time = datetime.now()
        client = self.get_client()
        
        if not client:
            response = self.get_fallback_response(message, document_context, filename)
            processing_time = (datetime.now() - start_time).total_seconds()
            return response, processing_time, len(document_context), chunk_count
        
        # Construir prompt otimizado com contexto do documento
        prompt_parts = [self.ipt_context]
        
        if document_context:
            prompt_parts.append(f"""
DOCUMENTO ENVIADO PELO USUÁRIO: {filename}
{'-' * 60}
{document_context}
{'-' * 60}

Com base no documento acima, responda à pergunta do usuário de forma precisa e contextualizada.
Mencione especificamente informações do documento quando relevante.
""")
        
        prompt_parts.append(f"\nPERGUNTA DO USUÁRIO: {message}\n\nResponda de forma clara, útil e baseada no contexto disponível.")
        
        prompt = "\n".join(prompt_parts)
        
        try:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=1500,
                messages=[{"role": "user", "content": prompt}]
            )
            
            processing_time = (datetime.now() - start_time).total_seconds()
            return response.content[0].text, processing_time, len(document_context), chunk_count
            
        except Exception as e:
            print(f"Erro na API Anthropic: {e}")
            processing_time = (datetime.now() - start_time).total_seconds()
            return self.get_fallback_response(message, document_context, filename), processing_time, len(document_context), chunk_count
    
    def get_fallback_response(self, message, document_context="", filename=""):
        message_lower = message.lower()
        
        if document_context:
            return f"Recebi e processei o documento '{filename}' com sucesso usando o sistema RAG avançado do Supabase. O conteúdo foi analisado e está disponível para consultas contextuais. Posso agora responder perguntas específicas sobre o documento, fazer análises detalhadas ou ajudar com interpretações baseadas no conteúdo. O que gostaria de saber?"
        
        if any(word in message_lower for word in ['o que é', 'quem é', 'como funciona', 'lidia']):
            return "Entendido, sou a assistente LIDIA do IPT (Instituto de Pesquisas Tecnológicas de São Paulo) com sistema RAG avançado integrado ao Supabase. Sou uma assistente conversacional de última geração, criada para apoiar colaboradores do IPT com orientações sobre pesquisa, redação técnica, análise avançada de documentos e metodologias. Como posso ajudá-lo hoje?"
        
        if any(word in message_lower for word in ['dados', 'análise', 'resultados']):
            return "Posso ajudá-lo a interpretar dados e discutir abordagens analíticas avançadas. Se você enviar documentos com seus dados ou resultados, posso fazer análises contextualizadas detalhadas usando meu sistema RAG. Que tipo de análise você precisa?"
        
        if any(word in message_lower for word in ['pesquisa', 'metodologia']):
            return "Posso auxiliá-lo na estruturação de metodologias de pesquisa e organização de ideias. Se você enviar documentos relacionados ao seu projeto, posso dar sugestões muito específicas e contextualizadas baseadas no conteúdo. Qual é o foco da sua pesquisa?"
        
        return "Olá! Sou a LIDIA, sua assistente inteligente do IPT com sistema RAG avançado integrado ao Supabase. Posso ajudá-lo com pesquisa, redação técnica, análise profunda de documentos e orientações metodológicas. Para respostas mais precisas e contextualizadas, você pode enviar documentos relevantes que analisarei automaticamente com algoritmos avançados. Como posso apoiá-lo hoje?"

# Inicializar componentes
db_manager = SupabaseManager()
assistant = LIDIAAssistant()
document_processor = AdvancedDocumentProcessor()

def get_current_costs():
    """Calcula custos baseado no Supabase"""
    if not db_manager.available:
        return {
            'total': 100,
            'fixed': 100,
            'variable': 0,
            'questions': 0,
            'budget_used': 50
        }
    
    try:
        # Usar dados do Supabase para cálculo mais preciso
        first_day = datetime.now().replace(day=1, hour=0, minute=0, second=0).isoformat()
        
        result = db_manager.client.table('conversation_logs').select('cost', count='exact').gte('timestamp', first_day).execute()
        
        questions_this_month = result.count if result.count else 0
        total_variable_cost = sum([float(row.get('cost', 0.003)) for row in result.data]) if result.data else 0
        
        fixed_costs = 100
        total_costs = fixed_costs + total_variable_cost
        budget_used = (total_costs / 200) * 100
        
        return {
            'total': total_costs,
            'fixed': fixed_costs,
            'variable': total_variable_cost,
            'questions': questions_this_month,
            'budget_used': budget_used
        }
    except Exception as e:
        print(f"[COSTS] Erro ao calcular custos: {e}")
        return {
            'total': 100,
            'fixed': 100,
            'variable': 0,
            'questions': 0,
            'budget_used': 50
        }

# Rotas principais
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin')
def admin_panel():
    if not session.get('authenticated') or not session.get('is_admin'):
        return redirect('/')
    return render_template('admin.html')

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email', '').strip()
    
    print(f"[AUTH] Tentativa de login: {email}")
    
    is_valid, message = db_manager.validate_ipt_email(email)
    
    if is_valid:
        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)
        user_agent = request.headers.get('User-Agent', '')
        db_manager.log_access(email, client_ip, "login", True, user_agent)
        
        is_admin = db_manager.is_admin(email)
        
        session['authenticated'] = True
        session['user_email'] = email
        session['is_admin'] = is_admin
        session['session_id'] = str(uuid.uuid4())
        
        print(f"[AUTH] Login sucesso - Email: {email}, Admin: {is_admin}")
        
        return jsonify({
            'success': True,
            'message': 'Login realizado com sucesso no sistema avançado',
            'is_admin': is_admin
        })
    else:
        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)
        user_agent = request.headers.get('User-Agent', '')
        db_manager.log_access(email, client_ip, "login_failed", False, user_agent)
        
        return jsonify({
            'success': False,
            'message': message
        }), 400

@app.route('/api/chat', methods=['POST'])
def chat():
    if not session.get('authenticated'):
        return jsonify({'error': 'Não autenticado'}), 401
    
    data = request.get_json()
    message = data.get('message', '').strip()
    chat_id = data.get('chat_id', '')
    
    if not message:
        return jsonify({'error': 'Mensagem vazia'}), 400
    
    if not chat_id:
        chat_id = 'chat_' + str(uuid.uuid4())
    
    try:
        # Verificar contexto de documento
        document_context, filename, chunk_count = db_manager.get_document_context(chat_id)
        
        response, processing_time, context_length, chunks_used = assistant.process_query(
            message, 
            session['user_email'],
            document_context or "",
            filename or "",
            chunk_count or 0
        )
        
        # Calcular custo melhorado
        base_cost = 0.003
        context_cost = (context_length / 1000) * 0.001
        total_cost = base_cost + context_cost + (processing_time * 0.0005)
        
        db_manager.log_conversation(
            session['user_email'], 
            chat_id,
            message, 
            response,
            cost=total_cost,
            document_context=document_context or "",
            response_time=processing_time,
            model=assistant.model,
            context_length=context_length,
            chunk_count=chunks_used
        )
        
        return jsonify({
            'response': response,
            'chat_id': chat_id,
            'timestamp': datetime.now().isoformat(),
            'processing_time': round(processing_time, 2),
            'context_used': bool(document_context),
            'context_length': context_length,
            'system_info': 'Powered by Supabase + Advanced RAG'
        })
        
    except Exception as e:
        print(f"[CHAT] Erro: {str(e)}")
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/api/chats')
def get_chats():
    if not session.get('authenticated'):
        return jsonify({'error': 'Não autenticado'}), 401
    
    chats = db_manager.get_user_chats(session['user_email'])
    return jsonify(chats)

@app.route('/api/chats/<chat_id>')
def get_chat(chat_id):
    if not session.get('authenticated'):
        return jsonify({'error': 'Não autenticado'}), 401
    
    messages = db_manager.get_chat_messages(session['user_email'], chat_id)
    return jsonify(messages)

@app.route('/api/document-context/<chat_id>')
def get_document_context_api(chat_id):
    if not session.get('authenticated'):
        return jsonify({'error': 'Não autenticado'}), 401
    
    document_context, filename, chunk_count = db_manager.get_document_context(chat_id)
    
    return jsonify({
        'has_document': document_context is not None,
        'filename': filename,
        'context_length': len(document_context) if document_context else 0,
        'chunk_count': chunk_count,
        'system': 'Supabase RAG'
    })

@app.route('/api/upload', methods=['POST'])
def upload_file():
    if not session.get('authenticated'):
        return jsonify({'error': 'Não autenticado'}), 401
    
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
    
    chat_id = request.form.get('chat_id', 'chat_' + str(uuid.uuid4()))
    
    try:
        start_time = time.time()
        
        print(f"[UPLOAD] Iniciando processamento avançado: {file.filename}")
        
        # Ler conteúdo original
        original_content = file.read()
        file.seek(0)
        
        # Processar com sistema avançado
        processed_text = document_processor.extract_text_from_file(file)
        processing_time = time.time() - start_time
        
        # Determinar método de extração
        extraction_method = "supabase_rag_advanced"
        if PYMUPDF_AVAILABLE and file.filename.lower().endswith('.pdf'):
            extraction_method += "_pymupdf"
        elif file.filename.lower().endswith('.docx'):
            extraction_method += "_docx2txt"
        
        # Estimar chunks
        estimated_chunks = max(1, len(processed_text.split('\n\n')))
        
        # Armazenar no Supabase
        success = db_manager.store_document(
            session['user_email'],
            chat_id,
            secure_filename(file.filename),
            original_content,
            processed_text,
            processing_time,
            extraction_method,
            estimated_chunks
        )
        
        if success:
            print(f"[UPLOAD] Sucesso: {file.filename} - {len(processed_text)} chars em {processing_time:.2f}s")
            
            return jsonify({
                'success': True,
                'filename': file.filename,
                'chat_id': chat_id,
                'message': f'Documento "{file.filename}" processado com sucesso usando sistema RAG avançado integrado ao Supabase! Conteúdo extraído, analisado e otimizado para consultas contextuais de alta performance.',
                'processing_stats': {
                    'original_size': len(original_content),
                    'processed_size': len(processed_text),
                    'processing_time': round(processing_time, 2),
                    'extraction_method': extraction_method,
                    'estimated_chunks': estimated_chunks,
                    'system': 'Supabase + Advanced RAG'
                }
            })
        else:
            return jsonify({'error': 'Erro ao armazenar documento no Supabase'}), 500
            
    except Exception as e:
        error_msg = f'Erro no processamento avançado: {str(e)}'
        print(f"[UPLOAD] {error_msg}")
        return jsonify({'error': error_msg}), 500

@app.route('/api/costs')
def costs():
    if not session.get('authenticated'):
        return jsonify({'error': 'Não autenticado'}), 401
    
    if not session.get('is_admin'):
        return jsonify({'error': 'Acesso negado - apenas administradores'}), 403
    
    costs_data = get_current_costs()
    return jsonify(costs_data)

# Rotas administrativas
@app.route('/api/admin/stats')
def admin_stats():
    if not session.get('authenticated') or not session.get('is_admin'):
        return jsonify({'error': 'Acesso negado'}), 403
    
    stats = db_manager.get_admin_stats()
    costs = get_current_costs()
    
    return jsonify({
        'stats': stats,
        'costs': costs,
        'system_info': {
            'database': 'Supabase PostgreSQL',
            'rag_system': 'Advanced RAG with Intelligent Chunking',
            'performance': 'High Performance + Auto Scaling'
        }
    })

@app.route('/api/admin/admins')
def get_admins():
    if not session.get('authenticated') or not session.get('is_admin'):
        return jsonify({'error': 'Acesso negado'}), 403
    
    admins = db_manager.get_admins()
    return jsonify(admins)

@app.route('/api/admin/admins', methods=['POST'])
def add_admin():
    if not session.get('authenticated') or not session.get('is_admin'):
        return jsonify({'error': 'Acesso negado'}), 403
    
    data = request.get_json()
    email = data.get('email', '').strip()
    
    if not email:
        return jsonify({'error': 'Email obrigatório'}), 400
    
    if not email.endswith('@ipt.br'):
        return jsonify({'error': 'Email deve ser do domínio @ipt.br'}), 400
    
    success, message = db_manager.add_admin(email, session['user_email'])
    
    if success:
        return jsonify({'message': message})
    else:
        return jsonify({'error': message}), 400

@app.route('/api/admin/admins/<email>', methods=['DELETE'])
def remove_admin(email):
    if not session.get('authenticated') or not session.get('is_admin'):
        return jsonify({'error': 'Acesso negado'}), 403
    
    success, message = db_manager.remove_admin(email)
    
    if success:
        return jsonify({'message': message})
    else:
        return jsonify({'error': message}), 400

@app.route('/api/logout', methods=['POST'])
def logout():
    if session.get('authenticated'):
        db_manager.log_access(
            session.get('user_email', ''),
            request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr),
            "logout",
            True
        )
    
    session.clear()
    return jsonify({'success': True})

@app.route('/api/user')
def user_info():
    if not session.get('authenticated'):
        return jsonify({'error': 'Não autenticado'}), 401
    
    user_data = {
        'email': session['user_email'],
        'is_admin': session.get('is_admin', False),
        'authenticated': True,
        'system': 'Supabase + Advanced RAG'
    }
    
    return jsonify(user_data)

@app.route('/health')
def health():
    return jsonify({
        'status': 'healthy', 
        'timestamp': datetime.now().isoformat(),
        'version': '4.0-supabase-rag-fixed',
        'database': 'Supabase PostgreSQL' if db_manager.available else 'Unavailable',
        'features': {
            'advanced_rag': True,
            'intelligent_chunking': True,
            'supabase_integration': db_manager.available,
            'pymupdf_available': PYMUPDF_AVAILABLE,
            'high_performance': True
        }
    })

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
